"""
Renders the fixed Tracker row skeleton (validated against the real DBIA
2026 Tracker sheet) into a formatted Tracker.xlsx, with the RFA's scoring
breakdown folded into the Competitive Check row as sub-bullets.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NUM_APPLICANT_COLUMNS = 8  # blank columns for the team to fill in per applicant

SECTION_HEADER_ROWS = {"Review & Submission Process"}
BOLD_ROWS = {"Farmer / Applicant"}


def _fixed_rows(reimbursement_summary: str, competitive_scoring_bullets: list[str],
                 grant_specific_rows: list[str]) -> list[str]:
    reimbursement_label = "Confirmed that this is a reimbursement style grant & applicant must pay up front / confirm match requirements"
    if reimbursement_summary:
        reimbursement_label += f"\n{reimbursement_summary}"

    competitive_label = "Competitive Check - Score for Evaluation Criteria (alignment with grant goals)"
    if competitive_scoring_bullets:
        bullets = "\n".join(f"• {b}" for b in competitive_scoring_bullets)
        competitive_label += f"\n{bullets}"

    rows = [
        "Farmer / Applicant",
        "Grant Writer",
        "Lasso CS",
        "Applicant information (e.g., location, EIN, etc.)",
        "Eligibility Check - yes/no",
        reimbursement_label,
        competitive_label,
        "Licensure Progress - yes/no",
        "Project data (e.g., goal, timeframe, total project cost, requested funds, previous awardee)",
        "Previous Funding",
        "Project summary",
        "Project workplan and timeline",
        "Project budget (including eligible and ineligible expenses)",
        "Budget justification",
        "Letter of commitment - financial institution",
        "Letters of support",
        "Vendor quotes",
    ]

    for extra_row in grant_specific_rows:
        rows.append(extra_row)

    rows += [
        "Review & Submission Process",
        "Peer review",
        "Applicant Review",
        "Internal team review (Molly/Nicole/Thalita)",
        "Application Submitted",
    ]

    return rows


def generate_tracker_xlsx(grant_name: str, year: str, tracker_extras: dict, output_path: Path) -> Path:
    reimbursement_summary = tracker_extras.get("reimbursement_summary", "")
    competitive_scoring_bullets = tracker_extras.get("competitive_scoring_bullets", [])
    grant_specific_rows = tracker_extras.get("grant_specific_rows", [])

    rows = _fixed_rows(reimbursement_summary, competitive_scoring_bullets, grant_specific_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tracker"
    ws["A1"] = f"{grant_name} ({year}) — Tracker"
    ws["A1"].font = Font(bold=True, size=14)

    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    start_row = 3
    for offset, label in enumerate(rows):
        row_idx = start_row + offset
        cell = ws.cell(row=row_idx, column=1, value=label)
        cell.alignment = wrap

        first_line = label.split("\n", 1)[0]
        if first_line in SECTION_HEADER_ROWS:
            cell.font = Font(bold=True)
            for col in range(1, NUM_APPLICANT_COLUMNS + 2):
                ws.cell(row=row_idx, column=col).fill = header_fill
        elif first_line in BOLD_ROWS:
            cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 55
    for col_idx in range(2, NUM_APPLICANT_COLUMNS + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
